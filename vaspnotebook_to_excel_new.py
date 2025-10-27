import os
import ase.io
import pandas as pd
from openpyxl import load_workbook

def collect_vasp_energies(base_dir, calcName_list, sheet_dir, output_filename="energies.xlsx", sheet_prefix="Energies"):
    """
    Collects final DFT energies from multiple VASP calculation folders and
    appends them as a new sheet in an existing Excel file (2nd, 3rd, etc. sheets).
    """

    results = []

    def process_folder(folder_name):
        """Read OUTCAR energy and append to results list."""
        try:
            folder_path = os.path.join(base_dir, folder_name)
            if not os.path.isdir(folder_path):
                print(f"⚠️ Folder not found: {folder_name}")
                return
            os.chdir(folder_path)
            atoms = ase.io.read("OUTCAR")
            energy = atoms.get_potential_energy()
            results.append((folder_name, energy))
        except Exception as e:
            print(f"⚠️ Error processing {folder_name}: {e}")

    # --- Process all folders ---
    for name in calcName_list:
        print(f"Processing {name} ...")
        process_folder(name)

    # --- Sort by energy ---
    results.sort(key=lambda x: x[1])
    df = pd.DataFrame(results, columns=["Folder", "Final Energy (eV)"])
    output_path = os.path.join(sheet_dir, output_filename)

    # --- Append to Excel file as a new sheet ---
    if os.path.exists(output_path):
        book = load_workbook(output_path)
        sheet_count = len(book.sheetnames)
        new_sheet_name = f"{sheet_prefix}_{sheet_count + 1}"  # e.g., Energies_2, Energies_3, etc.

        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name=new_sheet_name, index=False)

        print(f"✅ Energies written to new sheet '{new_sheet_name}' in {output_filename}")
    else:
        # If file doesn't exist, create it with the first sheet
        df.to_excel(output_path, index=False, sheet_name=f"{sheet_prefix}_1")
        print(f"📘 File not found. Created new Excel file with first sheet: {output_path}")

    print("\nSorted Energies:")
    print(df)
    return df, df.values.tolist()


# === Example usage ===
if __name__ == "__main__":
    base_dir = calcDir_
    sheet_dir = "/Users/toghrulazizli/Desktop/Research_/Olefin Metathesis/excelsheets"
    collect_vasp_energies(base_dir, calcName_list, sheet_dir, output_filename="MoO3_O_MO2.xlsx", sheet_name="Prop_H2_Modimer")

